from __future__ import annotations

import csv
import json
from dataclasses import dataclass, fields
from io import BytesIO, StringIO

from workload_analytics.dashboard.queries import DashboardData
from workload_analytics.models import DeveloperPeriodMetrics

_IDENTITY_FIELDS = ("granularity", "developer_email", "period_start", "period_end")
_EXPORT_FIELDS = _IDENTITY_FIELDS + tuple(
    f.name for f in fields(DeveloperPeriodMetrics) if f.name not in _IDENTITY_FIELDS
)


@dataclass(frozen=True, slots=True)
class ExportedCsv:
    file_name: str
    content: str


@dataclass(frozen=True, slots=True)
class ExportedJson:
    file_name: str
    content: str


@dataclass(frozen=True, slots=True)
class ExportedExcel:
    file_name: str
    content: bytes


def build_filtered_metrics_csv(data: DashboardData) -> ExportedCsv:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(list(_EXPORT_FIELDS))

    for row in data.filtered_metrics:
        writer.writerow(_metric_row_values(row))

    return ExportedCsv(
        file_name=_base_file_name(data) + ".csv",
        content=buffer.getvalue(),
    )


def build_filtered_metrics_json(data: DashboardData) -> ExportedJson:
    rows = [
        dict(zip(_EXPORT_FIELDS, _metric_row_values(row)))
        for row in data.filtered_metrics
    ]

    summary = data.summary
    payload = {
        "filters": {
            "start_date": data.filters.start_date.isoformat(),
            "end_date": data.filters.end_date.isoformat(),
            "granularity": data.filters.granularity.value,
            "developer_email": data.filters.developer_email,
        },
        "summary": {
            "active_developers": summary.active_developers,
            "github_prs_merged": summary.github_prs_merged,
            "github_commits_landed": summary.github_commits_landed,
            "github_lines_added": summary.github_lines_added,
            "github_lines_deleted": summary.github_lines_deleted,
            "jira_issues_assigned": summary.jira_issues_assigned,
            "successful_deployments": summary.successful_deployments,
            "failed_deployments": summary.failed_deployments,
        },
        "comparison": [
            {
                "developer_email": r.developer_email,
                "github_prs_merged": r.github_prs_merged,
                "github_commits_landed": r.github_commits_landed,
                "jira_issues_assigned": r.jira_issues_assigned,
            }
            for r in data.comparison_rows
        ],
        "metrics": rows,
    }

    return ExportedJson(
        file_name=_base_file_name(data) + ".json",
        content=json.dumps(payload, ensure_ascii=False, indent=2),
    )


def build_filtered_metrics_excel(data: DashboardData) -> ExportedExcel:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    wb = Workbook()

    # Sheet 1: Developer period metrics
    ws_metrics = wb.active
    ws_metrics.title = "개발자별 기간 지표"
    ws_metrics.append(list(_EXPORT_FIELDS))
    for row in data.filtered_metrics:
        ws_metrics.append(list(_metric_row_values(row)))

    # Sheet 2: Team summary
    ws_summary = wb.create_sheet("팀 요약")
    summary = data.summary
    ws_summary.append(["지표", "값"])
    for label, value in [
        ("활성 개발자", summary.active_developers),
        ("머지된 PR", summary.github_prs_merged),
        ("랜딩된 커밋", summary.github_commits_landed),
        ("추가된 라인", summary.github_lines_added),
        ("삭제된 라인", summary.github_lines_deleted),
        ("할당된 이슈", summary.jira_issues_assigned),
        ("성공 배포", summary.successful_deployments),
        ("실패 배포", summary.failed_deployments),
        ("Stale PR", summary.github_prs_stale),
    ]:
        ws_summary.append([label, value])

    # Sheet 3: Developer comparison
    ws_compare = wb.create_sheet("개발자별 비교")
    ws_compare.append([
        "developer_email", "github_prs_merged", "github_commits_landed",
        "github_lines_added", "github_lines_deleted", "jira_issues_assigned",
    ])
    for r in data.comparison_rows:
        ws_compare.append([
            r.developer_email, r.github_prs_merged, r.github_commits_landed,
            r.github_lines_added, r.github_lines_deleted, r.jira_issues_assigned,
        ])

    # Sheet 4: Delivery metrics
    ws_delivery = wb.create_sheet("배포 지표")
    ws_delivery.append([
        "period_start", "period_end",
        "successful_deployments", "failed_deployments",
        "deployment_lead_time_hours", "deployments_with_lead_time",
    ])
    for d in data.delivery_trend_points:
        ws_delivery.append([
            d.period_start.isoformat(), d.period_end.isoformat(),
            d.successful_deployments, d.failed_deployments,
            d.deployment_lead_time_hours, d.deployments_with_lead_time,
        ])

    buffer = BytesIO()
    wb.save(buffer)
    return ExportedExcel(
        file_name=_base_file_name(data) + ".xlsx",
        content=buffer.getvalue(),
    )


def _metric_row_values(row: DeveloperPeriodMetrics) -> tuple[object, ...]:
    return (
        row.granularity.value,
        row.developer_email,
        row.period_start.isoformat(),
        row.period_end.isoformat(),
        *(getattr(row, field) for field in _EXPORT_FIELDS[4:]),
    )


def _base_file_name(data: DashboardData) -> str:
    return (
        "workload-analytics_"
        f"{data.filters.granularity.value}_"
        f"{data.filters.start_date.isoformat()}_"
        f"{data.filters.end_date.isoformat()}"
    )
